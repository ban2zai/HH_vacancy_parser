import requests
import json
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import re
import openpyxl
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse

def format_date_time_separate(iso_date_string):
    if iso_date_string is None or iso_date_string == "Не найдено" or iso_date_string == "Автообновление не настроено":
        return "", ""
    try:
        if '.' in iso_date_string:
            iso_date_string = iso_date_string.split('.')[0]
        if '+' in iso_date_string:
            iso_date_string = iso_date_string.split('+')[0]
        elif 'Z' in iso_date_string:
            iso_date_string = iso_date_string.replace('Z', '')

        dt_object = datetime.strptime(iso_date_string, "%Y-%m-%dT%H:%M:%S")
        return dt_object.strftime("%d.%m.%Y"), dt_object.strftime("%H:%M")
    except ValueError:
        print(f"Предупреждение: Не удалось распарсить дату/время '{iso_date_string}'.")
        return iso_date_string, ""

def fetch_specializations_from_api(api_url: str) -> dict:
    print(f"Получение справочника специализаций с {api_url}...")
    try:
        response = requests.get(api_url, timeout=10)
        response.raise_for_status()
        data = response.json()
        print("Справочник специализаций успешно получен.")
        return data
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при получении справочника специализаций с API: {e}")
        return {}
    except json.JSONDecodeError as e:
        print(f"Ошибка декодирования JSON при получении справочника специализаций: {e}")
        return {}

def create_specialization_lookup_table(json_data: dict) -> dict:
    lookup_table = {}
    if "categories" in json_data and isinstance(json_data["categories"], list):
        for category in json_data["categories"]:
            if "roles" in category and isinstance(category["roles"], list):
                for role in category["roles"]:
                    if "id" in role and "name" in role:
                        lookup_table[str(role["id"])] = role["name"]
    return lookup_table

def parse_hh_vacancies(base_url: str, max_pages: int, specialization_lookup: dict = None) -> list[dict]:
    all_vacancies_data = []
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    vacancy_counter = 0
    max_pages = 20 if max_pages == 0 or max_pages > 20 else max_pages
    seen_vacancy_links = set()  # Множество для отслеживания уникальных ссылок

    if "&items_on_page=" not in base_url:
        base_url += "&items_on_page=100"
    else:
        base_url = re.sub(r"&items_on_page=\d+", "&items_on_page=100", base_url)

    for current_page in range(max_pages):
        page_url = f"{base_url}&page={current_page}"
        print(f"Загрузка страницы {current_page + 1}: {page_url}")

        try:
            response = requests.get(page_url, headers=headers, timeout=15)
            response.raise_for_status()
            html_content = response.text

            soup = BeautifulSoup(html_content, 'html.parser')
            initial_state_tag = soup.find('template', id='HH-Lux-InitialState')

            if not initial_state_tag:
                print(f"Ошибка: Тег <template id='HH-Lux-InitialState'> не найден на странице {current_page + 1}.")
                if current_page == 0:
                    return []
                break

            json_data_str = initial_state_tag.string
            if not json_data_str:
                print(
                    f"Ошибка: JSON-строка внутри тега <template id='HH-Lux-InitialState'> пуста на странице {current_page + 1}.")
                if current_page == 0:
                    return []
                break

            data = json.loads(json_data_str)

            current_page_vacancies = []
            if "vacancySearchResult" in data and isinstance(data["vacancySearchResult"], dict):
                if "vacancies" in data["vacancySearchResult"] and \
                        isinstance(data["vacancySearchResult"]["vacancies"], list):
                    current_page_vacancies = data["vacancySearchResult"]["vacancies"]
                else:
                    print(
                        f"В 'vacancySearchResult' не найден массив 'vacancies' или он имеет неверный формат на странице {current_page + 1}.")
            else:
                print(
                    f"В JSON-данных не найден ключ 'vacancySearchResult' или он имеет неверный формат на странице {current_page + 1}.")

            if not current_page_vacancies:
                print(f"На странице {current_page + 1} вакансий не найдено. Завершение парсинга.")
                break

            for vacancy in current_page_vacancies:
                vacancy_id = vacancy.get("vacancyId")
                vacancy_link = f"https://hh.ru/vacancy/{vacancy_id}" if vacancy_id else None

                # Пропускаем вакансию, если ссылка уже встречалась
                if vacancy_link and vacancy_link in seen_vacancy_links:
                    print(f"Вакансия с ссылкой {vacancy_link} уже обработана, пропускаем.")
                    continue

                vacancy_counter += 1
                if vacancy_link:
                    seen_vacancy_links.add(vacancy_link)

                city = vacancy.get("area", {}).get("name")
                job_title = vacancy.get("name")
                company_name = vacancy.get("company", {}).get("name")

                compencation_mode = vacancy.get("compensation", {}).get("mode")
                compencation_currencyCode = vacancy.get("compensation", {}).get("currencyCode")
                if compencation_mode == 'MONTH' and compencation_currencyCode == 'RUR':
                    compensation_from = vacancy.get("compensation", {}).get("from")
                    compensation_to = vacancy.get("compensation", {}).get("to")
                else:
                    compensation_from = None
                    compensation_to = None

                work_experience = vacancy.get("workExperience")

                publication_type = "Неизвестно"
                calculated_states = vacancy.get("vacancyProperties", {}).get("calculatedStates", {}).get("HH", {})
                premium = calculated_states.get("premium", False)
                standard = calculated_states.get("standard", False)
                standard_plus = calculated_states.get("standardPlus", False)

                if premium:
                    publication_type = "Премиум"
                elif standard:
                    publication_type = "Стандарт"
                elif standard_plus:
                    publication_type = "Плюс Стандарт"

                creation_time_raw = vacancy.get("creationTime")
                creation_date_formatted, creation_time_formatted = format_date_time_separate(creation_time_raw)

                publication_date_obj = vacancy.get("publicationTime", {})
                publication_date_raw = publication_date_obj.get("$") if publication_date_obj else None
                publication_date_formatted, publication_time_formatted = format_date_time_separate(publication_date_raw)

                total_responses_count = vacancy.get("totalResponsesCount")
                responses_count = vacancy.get("responsesCount")

                is_adv = vacancy.get("@isAdv", False)
                has_hh_auction = False
                has_zp_promo = False
                click_url = vacancy.get("clickUrl", None)

                vacancy_properties = vacancy.get("vacancyProperties", {})
                if "properties" in vacancy_properties and isinstance(vacancy_properties["properties"], list):
                    for prop_group in vacancy_properties["properties"]:
                        if "property" in prop_group and isinstance(prop_group["property"], list):
                            for prop in prop_group["property"]:
                                if prop.get("propertyType") == "HH_AUCTION":
                                    has_hh_auction = True
                                if prop.get("propertyType") == "ZP_PROMO":
                                    has_zp_promo = True

                professional_role_id = None
                professional_role_name = "Неизвестно"
                if "professionalRoleIds" in vacancy and isinstance(vacancy["professionalRoleIds"], list) and \
                        len(vacancy["professionalRoleIds"]) > 0 and \
                        "professionalRoleId" in vacancy["professionalRoleIds"][0] and \
                        isinstance(vacancy["professionalRoleIds"][0]["professionalRoleId"], list) and \
                        len(vacancy["professionalRoleIds"][0]["professionalRoleId"]) > 0:
                    professional_role_id = str(vacancy["professionalRoleIds"][0]["professionalRoleId"][0])
                    if specialization_lookup and professional_role_id in specialization_lookup:
                        professional_role_name = specialization_lookup[professional_role_id]
                    else:
                        professional_role_name = f"ID: {professional_role_id}"

                hours_since_creation = None
                if creation_date_formatted and creation_time_formatted:
                    try:
                        creation_datetime_str = f"{creation_date_formatted} {creation_time_formatted}"
                        creation_datetime_obj = datetime.strptime(creation_datetime_str, "%d.%m.%Y %H:%M")
                        current_datetime = datetime.now()
                        time_difference = current_datetime - creation_datetime_obj
                        hours_since_creation = round((time_difference.total_seconds() / 3600) / 24, 2)
                    except ValueError:
                        hours_since_creation = "Ошибка даты/времени"

                all_vacancies_data.append({
                    "№ ": vacancy_counter,
                    "Город   ": city,
                    "Вакансия": job_title,
                    "Опыт работы": work_experience,
                    "Компания": company_name,
                    "Ссылка": vacancy_link,
                    "Тип публикации": publication_type,
                    "isAdv": "Да" if is_adv else "Нет",
                    "HH_AUCTION": "Да" if has_hh_auction else "Нет",
                    "ЗП От": compensation_from,
                    "ЗП До": compensation_to,
                    "Дата соз": creation_date_formatted,
                    "Время с": creation_time_formatted,
                    "Дата пуб": publication_date_formatted,
                    "Время п": publication_time_formatted,
                    "О(общ)": total_responses_count,
                    "О(2)": responses_count,
                    "Специализация": professional_role_name,
                    "Дней Прошло": hours_since_creation
                })

        except requests.exceptions.Timeout:
            print(f"Ошибка: Превышено время ожидания запроса для URL: {page_url}. Завершение парсинга.")
            break
        except requests.exceptions.RequestException as e:
            print(f"Ошибка при запросе к URL {page_url}: {e}. Пропуск страницы.")
            continue
        except json.JSONDecodeError as e:
            print(f"Ошибка при декодировании JSON на странице {current_page + 1}: {e}. Пропуск страницы.")
            continue
        except Exception as e:
            print(f"Произошла непредвиденная ошибка на странице {current_page + 1}: {e}. Пропуск страницы.")
            continue

    return all_vacancies_data

def save_to_excel(data: list[dict]):
    if not data:
        print("Нет данных для сохранения в Excel.")
        return

    df = pd.DataFrame(data)
    df["Откликов в день"] = ""

    first_city = "НеизвестныйГород"
    for row in data:
        if "Город   " in row and row["Город   "]:
            first_city = re.sub(r'[\\/:*?"<>|]', '_', row["Город   "])
            break

    current_date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_name = f"{first_city}_{current_date}.xlsx"

    try:
        df.to_excel(file_name, index=False, engine='openpyxl')
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active

        header_row = ws[1]
        total_responses_col = None
        days_passed_col = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == "О(общ)":
                total_responses_col = get_column_letter(idx)
            if cell.value == "Дней Прошло":
                days_passed_col = get_column_letter(idx)

        if total_responses_col and days_passed_col:
            responses_per_day_col = get_column_letter(ws.max_column)
            for row in range(2, ws.max_row + 1):
                ws[f"{responses_per_day_col}{row}"] = f'=IFERROR(ROUND({total_responses_col}{row}/{days_passed_col}{row},2),"")'

        wb.save(file_name)
        print(f"\nДанные успешно сохранены в файл: {file_name}")
    except Exception as e:
        print(f"Ошибка при сохранении в Excel: {e}")

def is_valid_hh_url(url: str) -> bool:
    """Проверяет, является ли URL корректным для hh.ru поиска вакансий."""
    if not url:
        return False
    try:
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        parsed_url = urlparse(url)
        if not parsed_url.netloc.endswith('hh.ru'):
            return False
        if 'search/vacancy' not in parsed_url.path:
            return False
        return True
    except Exception:
        return False

if __name__ == "__main__":
    specialization_api_url = "https://api.hh.ru/professional_roles"
    specialization_json_data = fetch_specializations_from_api(specialization_api_url)

    specializations_map = {}
    if specialization_json_data:
        specializations_map = create_specialization_lookup_table(specialization_json_data)
    else:
        print(
            "Не удалось получить справочник специализаций с API. Названия специализаций будут отображаться как 'ID: [числовой ID]'.")

    while True:
        base_url_input = input("Введите базовый URL для парсинга: ").strip()
        if is_valid_hh_url(base_url_input):
            if not base_url_input.startswith(('http://', 'https://')):
                base_url_input = 'https://' + base_url_input
            break
        print("Ошибка: Введён пустой или неверный URL. URL должен содержать 'search/vacancy' и домен, заканчивающийся на 'hh.ru'.")

    while True:
        num_pages_input = input("Введите количество страниц для парсинга (0 - все страницы, но не более 20): ").strip()
        if num_pages_input == "":
            print("Ошибка: Поле не может быть пустым. Попробуйте снова.")
            continue
        try:
            num_pages_to_parse = int(num_pages_input)
            if num_pages_to_parse < 0:
                print("Ошибка: Количество страниц не может быть отрицательным. Попробуйте снова.")
                continue
            break
        except ValueError:
            print("Ошибка: Введите число от 0 до 20.")

    if num_pages_to_parse == 0 or num_pages_to_parse > 20:
        print("Количество страниц для парсинга: 20")
        num_pages_to_parse = 20

    all_extracted_data = parse_hh_vacancies(base_url_input, num_pages_to_parse, specializations_map)
    save_to_excel(all_extracted_data)

    print("\nПарсинг завершен.")